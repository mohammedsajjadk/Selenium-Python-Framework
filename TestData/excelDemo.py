import openpyxl

book = openpyxl.load_workbook("PythonDemo.xlsx")
sheet = book.active
Dict = {}  # storing in dictionary
cell = sheet.cell(row=1, column=2)
print(cell.value)  # Output: firstName
sheet.cell(row=2, column=2).value = "Sajjad"  # writing
print(sheet.cell(row=2, column=2).value)  # reading the value we have written  # Output: Sajjad
print(sheet.max_row)  # Output: 3
print(sheet.max_column)  # Output: 4
print(sheet['A1'].value)  # Output: Name

'''Iterating rows and columns and extracting only testcase2 row values'''
for i in range(1, sheet.max_row + 1):  # to get rows
    if sheet.cell(row=i, column=1).value == "Testcase2":  # checking first column value is TestCase2

        for j in range(2, sheet.max_column + 1):  # to get columns
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)  # Output: {'firstName': 'Farheen', 'lastName': 'Taj', 'email': 'far@gmail.com'}
