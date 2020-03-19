import openpyxl


class HomePageData:
    test_HomePage_data = [{"first_name": "Sajjad", "last_name": "K", "gender": "Male"},
                          {"first_name": "Farheen", "last_name": "M", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook(
            "I:\Pessoal\Materials\Python and Selenium\Selenium with Python\Projects\PythonSeleniumFramework\TestData\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]  # sending as list because params accept list
